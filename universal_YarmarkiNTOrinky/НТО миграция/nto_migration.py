from __future__ import annotations

import argparse
import copy
import json
import os
import traceback
import warnings
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning

from _api import (
    build_file_meta,
    create_record,
    delete_record,
    get_runtime_base_url,
    get_runtime_ui_base_url,
    set_runtime_urls,
    setup_session,
    update_record,
    upload_file,
)
from _config import (
    BASE_URL,
    EXCEL_FILE_NAME,
    EXCEL_INPUT_GLOB,
    FILES_DIR,
    JWT_URL,
    NTO_MESTO_COLLECTION,
    NTO_NSI_COLLECTION,
    RESUME_BY_DEFAULT,
    ROLLBACK_BODY_FILE,
    SCRIPT_DIR,
    SHEET_MESTO,
    SHEET_TORGI,
    STATE_FILE,
    TORGI_COLLECTION,
    UI_BASE_URL,
)
from _excel_input import discover_excel_files
from _logger import setup_fail_logger, setup_logger, setup_success_logger, setup_user_logger
from _profiles import PROFILES
from _state import ResumeState
from _utils import dump_json, parse_key_value_mapping, parse_path_list, resolve_local_file_path, set_by_path
from _vba_nto import (
    PendingUpload,
    build_nsi_local_object_nto_payload,
    iter_vba_rows,
    norm_str,
    transform_row_to_bidding,
    transform_row_to_registry,
)


@dataclass
class WorkbookRunSpec:
    workbook_path: str
    files_dir: str


def _console_block(title: str, lines: Optional[List[str]] = None, width: int = 92) -> str:
    safe_title = str(title or "").strip() or "Инфо"
    content = ["", "=" * width, safe_title, "-" * width]
    for line in (lines or []):
        content.append(str(line))
    content.append("=" * width)
    return "\n".join(content)


def _format_iso_for_console(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return "-"
    try:
        normalized = text[:-1] + "+00:00" if text.endswith("Z") else text
        dt = datetime.fromisoformat(normalized)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone().strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return text


def _build_file_placeholder(*, filename: str, size: int, entity_field_path: str, allow_external: bool) -> Dict[str, Any]:
    return {
        "_id": "",
        "originalName": filename,
        "size": int(size or 0),
        "isFile": True,
        "entityFieldPath": entity_field_path,
        "allowExternal": bool(allow_external),
    }


def _rollback_payload(rollback_items: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {
            "_id": item.get("_id"),
            "guid": item.get("guid"),
            "parentEntries": item.get("parentEntries"),
        }
        for item in rollback_items
    ]


def _append_unique_rollback(target: List[Dict[str, Any]], entry: Dict[str, Any]) -> None:
    key = (str(entry.get("parentEntries")), str(entry.get("_id")))
    for existing in target:
        if (str(existing.get("parentEntries")), str(existing.get("_id"))) == key:
            return
    target.append(entry)


def _write_rollback_body(rollback_items: Sequence[Dict[str, Any]]) -> None:
    dump_json(Path(ROLLBACK_BODY_FILE), _rollback_payload(list(rollback_items)))


def _resolve_uploads(
    uploads: Sequence[PendingUpload],
    *,
    files_dir: Path,
    base_dir: Path,
) -> List[Dict[str, Any]]:
    resolved: List[Dict[str, Any]] = []
    for task in uploads:
        local_path = resolve_local_file_path(task.filename, files_dir=files_dir, base_dir=base_dir)
        if local_path:
            filename = os.path.basename(local_path)
            size = os.path.getsize(local_path)
        elif task.allow_external:
            filename = os.path.basename(str(task.filename or "")) or str(task.filename or "")
            size = 0
        else:
            raise FileNotFoundError(f"Файл для загрузки не найден: {task.filename}")
        resolved.append(
            {
                "task": task,
                "local_path": local_path,
                "filename": filename,
                "size": size,
            }
        )
    return resolved


def _apply_pending_uploads(
    *,
    session,
    logger,
    collection: str,
    main_id: str,
    guid: str,
    doc_state: Dict[str, Any],
    uploads: Sequence[PendingUpload],
    files_dir: Path,
    base_dir: Path,
    continue_on_upload_error: bool = False,
) -> Dict[str, Any]:
    if not uploads:
        return doc_state

    resolved = _resolve_uploads(uploads, files_dir=files_dir, base_dir=base_dir)

    for item in resolved:
        task = item["task"]
        set_by_path(
            doc_state,
            task.target_path,
            _build_file_placeholder(
                filename=item["filename"],
                size=item["size"],
                entity_field_path=task.target_path,
                allow_external=task.allow_external,
            ),
        )

    update_record(session, logger, collection=collection, main_id=main_id, guid=guid, body=doc_state)

    for item in resolved:
        task = item["task"]
        local_path = item["local_path"]
        if not local_path:
            logger.info(
                "[UPLOAD][SKIP-EXTERNAL] collection=%s id=%s field=%s file=%s",
                collection,
                main_id,
                task.target_path,
                task.filename,
            )
            continue
        try:
            uploaded = upload_file(
                session,
                logger,
                entry_name=collection,
                entry_id=main_id,
                entity_field_path=task.target_path,
                file_path=local_path,
                allow_external=task.allow_external,
            )
            meta = build_file_meta(
                uploaded,
                fallback_name=item["filename"],
                fallback_size=item["size"],
                path=task.target_path,
                allow_external=task.allow_external,
            )
            set_by_path(doc_state, task.target_path, meta)
        except Exception:
            logger.exception(
                "[UPLOAD][FAIL] collection=%s id=%s field=%s file=%s",
                collection,
                main_id,
                task.target_path,
                local_path,
            )
            if not continue_on_upload_error:
                raise

    updated = update_record(session, logger, collection=collection, main_id=main_id, guid=guid, body=doc_state)
    if isinstance(updated, dict):
        return updated
    return doc_state


def _rollback_created_records(session, logger, created_refs: Sequence[Dict[str, str]]) -> None:
    for item in reversed(list(created_refs)):
        collection = str(item.get("collection") or "").strip()
        main_id = str(item.get("_id") or "").strip()
        guid = str(item.get("guid") or "").strip()
        if not collection or not main_id:
            continue
        try:
            status = delete_record(session, logger, collection=collection, main_id=main_id, guid=guid)
            logger.warning("[ROLLBACK-ROW] deleted %s/%s -> HTTP %s", collection, main_id, status)
        except Exception:
            logger.exception("[ROLLBACK-ROW] failed to delete %s/%s", collection, main_id)


def _process_mesto_row(
    *,
    row_idx: int,
    row_data: Dict[str, Any],
    session,
    logger,
    success_logger,
    rollback_items: List[Dict[str, Any]],
    files_dir: Path,
    base_dir: Path,
    dry_run: bool,
) -> Dict[str, Any]:
    payload, pending_uploads = transform_row_to_registry(row_data, session, logger)
    if dry_run:
        logger.info("[DRY][%s][ROW:%s] payload keys=%s uploads=%s", SHEET_MESTO, row_idx, sorted(payload.keys()), len(pending_uploads))
        return {"collection": NTO_MESTO_COLLECTION, "_id": "", "guid": str(payload.get("guid") or "")}

    created_refs: List[Dict[str, str]] = []
    entries_to_log: List[Dict[str, Any]] = []
    try:
        created = create_record(session, logger, NTO_MESTO_COLLECTION, payload)
        main_id = str(created.get("_id") or "")
        final_guid = str(created.get("guid") or payload.get("guid") or "")
        if not main_id or not final_guid:
            raise RuntimeError("Нет _id/guid у созданной записи NTOmesto")

        created_refs.append({"collection": NTO_MESTO_COLLECTION, "_id": main_id, "guid": final_guid})

        payload["guid"] = final_guid
        gos_link = f"https://www.gosuslugi.ru/trade/{final_guid}"
        payload.setdefault("ntoInformation", {})
        payload["ntoInformation"]["GosuslugiData"] = gos_link

        doc_state = {"_id": main_id, "guid": final_guid, **copy.deepcopy(payload)}
        doc_state["guid"] = final_guid
        doc_state.setdefault("ntoInformation", {})
        doc_state["ntoInformation"]["GosuslugiData"] = gos_link

        project_status_name = norm_str(((payload.get("ntoInformation") or {}).get("ProjectStatus") or {}).get("name")).lower()
        if project_status_name == "утверждено":
            nsi_payload = build_nsi_local_object_nto_payload({**payload, "guid": final_guid})
            created_nsi = create_record(session, logger, NTO_NSI_COLLECTION, nsi_payload)
            nsi_id = str(created_nsi.get("_id") or "")
            nsi_guid = str(created_nsi.get("guid") or nsi_payload.get("guid") or "")
            if not nsi_id:
                raise RuntimeError("Нет _id у созданной записи nsiLocalObjectNTO")
            created_refs.append({"collection": NTO_NSI_COLLECTION, "_id": nsi_id, "guid": nsi_guid})
            entries_to_log.append(
                {
                    "sheet": SHEET_MESTO,
                    "row": row_idx,
                    "_id": nsi_id,
                    "guid": nsi_guid,
                    "parentEntries": NTO_NSI_COLLECTION,
                }
            )

        if any(task.target_path == "blockDopSogl[0].SupplementalAgreementIfAny" for task in pending_uploads):
            if not isinstance(doc_state.get("blockDopSogl"), list):
                doc_state["blockDopSogl"] = []
            while len(doc_state["blockDopSogl"]) == 0:
                doc_state["blockDopSogl"].append({})

        if pending_uploads:
            _apply_pending_uploads(
                session=session,
                logger=logger,
                collection=NTO_MESTO_COLLECTION,
                main_id=main_id,
                guid=final_guid,
                doc_state=doc_state,
                uploads=pending_uploads,
                files_dir=files_dir,
                base_dir=base_dir,
                continue_on_upload_error=False,
            )

        entries_to_log.append(
            {
                "sheet": SHEET_MESTO,
                "row": row_idx,
                "_id": main_id,
                "guid": final_guid,
                "parentEntries": NTO_MESTO_COLLECTION,
                "uploads": len(pending_uploads),
            }
        )

        for entry in entries_to_log:
            success_logger.info(json.dumps(entry, ensure_ascii=False))
            _append_unique_rollback(rollback_items, entry)
        _write_rollback_body(rollback_items)
        return {"collection": NTO_MESTO_COLLECTION, "_id": main_id, "guid": final_guid}
    except Exception:
        _rollback_created_records(session, logger, created_refs)
        raise


def _process_torgi_row(
    *,
    row_idx: int,
    row_data: Dict[str, Any],
    session,
    logger,
    success_logger,
    rollback_items: List[Dict[str, Any]],
    files_dir: Path,
    base_dir: Path,
    dry_run: bool,
) -> Dict[str, Any]:
    payload, pending_uploads = transform_row_to_bidding(row_data, session, logger)
    if dry_run:
        logger.info("[DRY][%s][ROW:%s] payload keys=%s uploads=%s", SHEET_TORGI, row_idx, sorted(payload.keys()), len(pending_uploads))
        return {"collection": TORGI_COLLECTION, "_id": "", "guid": str(payload.get("guid") or "")}

    created_refs: List[Dict[str, str]] = []
    try:
        created = create_record(session, logger, TORGI_COLLECTION, payload)
        main_id = str(created.get("_id") or "")
        guid = str(created.get("guid") or payload.get("guid") or "")
        if not main_id or not guid:
            raise RuntimeError("Нет _id/guid у созданной записи reestrbiddingReestr")

        created_refs.append({"collection": TORGI_COLLECTION, "_id": main_id, "guid": guid})
        auid = created.get("auid")
        payload["guid"] = guid
        doc_state = {"_id": main_id, "guid": guid, **({"auid": auid} if auid else {}), **copy.deepcopy(payload)}

        if pending_uploads:
            _apply_pending_uploads(
                session=session,
                logger=logger,
                collection=TORGI_COLLECTION,
                main_id=main_id,
                guid=guid,
                doc_state=doc_state,
                uploads=pending_uploads,
                files_dir=files_dir,
                base_dir=base_dir,
                continue_on_upload_error=False,
            )

        entry = {
            "sheet": SHEET_TORGI,
            "row": row_idx,
            "_id": main_id,
            "guid": guid,
            "parentEntries": TORGI_COLLECTION,
            "uploads": len(pending_uploads),
        }
        success_logger.info(json.dumps(entry, ensure_ascii=False))
        _append_unique_rollback(rollback_items, entry)
        _write_rollback_body(rollback_items)
        return {"collection": TORGI_COLLECTION, "_id": main_id, "guid": guid}
    except Exception:
        _rollback_created_records(session, logger, created_refs)
        raise


def _iter_selected_sheets(selected: str) -> List[Tuple[str, str]]:
    mapping = []
    if selected in {"all", "mesto"}:
        mapping.append((SHEET_MESTO, "mesto"))
    if selected in {"all", "torgi"}:
        mapping.append((SHEET_TORGI, "torgi"))
    return mapping


def _prompt_with_default(label: str, default_value: str, interactive: bool) -> str:
    if not interactive:
        return default_value
    entered = input(f"{label} [{default_value}]: ").strip().strip('"').strip("'")
    return entered or default_value


def _ask_error_action() -> str:
    print("Ошибка строки. Действие: [r]etry / [s]kip / [a]bort")
    while True:
        answer = input("Выбор: ").strip().lower()
        if answer in {"r", "retry"}:
            return "retry"
        if answer in {"s", "skip"}:
            return "skip"
        if answer in {"a", "abort", "stop"}:
            return "abort"
        if not answer:
            return "abort"
        print("Введите r, s или a")


def _numeric_hints(text: str) -> List[str]:
    normalized = (" " + str(text or "").strip().lower() + " ")
    hints = set()
    for token in normalized.split():
        if token.isdigit():
            hints.add(token)
    word_map = {
        " one ": "1",
        " first ": "1",
        " two ": "2",
        " second ": "2",
        " three ": "3",
        " third ": "3",
        " один ": "1",
        " первый ": "1",
        " первая ": "1",
        " два ": "2",
        " второй ": "2",
        " вторая ": "2",
        " три ": "3",
        " третий ": "3",
        " третья ": "3",
    }
    for needle, number in word_map.items():
        if needle in normalized:
            hints.add(number)
    return sorted(hints)


def _choose_single_workbook(candidates: List[str], interactive: bool) -> Optional[str]:
    if not candidates:
        return None
    if len(candidates) == 1 or not interactive:
        return candidates[0]
    print("\nДоступные книги для миграции:")
    for i, wb in enumerate(candidates, start=1):
        print(f"  {i}) {wb}")
    raw = input("Выберите номер книги [1]: ").strip()
    if not raw:
        return candidates[0]
    try:
        idx = int(raw)
    except Exception:
        return candidates[0]
    if 1 <= idx <= len(candidates):
        return candidates[idx - 1]
    return candidates[0]


def _choose_mass_workbooks(candidates: List[str], interactive: bool) -> List[str]:
    if not candidates:
        return []
    if not interactive:
        return list(candidates)
    print("\nДоступные книги для миграции:")
    for i, wb in enumerate(candidates, start=1):
        print(f"  {i}) {wb}")
    raw = input("Введите номера книг через запятую или Enter для всех: ").strip()
    if not raw:
        return list(candidates)
    selected: List[str] = []
    for part in raw.split(","):
        token = part.strip()
        if not token:
            continue
        try:
            idx = int(token)
        except Exception:
            continue
        if 1 <= idx <= len(candidates):
            selected.append(candidates[idx - 1])
    return selected or list(candidates)


def _resolve_mode(mode: str, candidates: List[str], interactive: bool) -> str:
    if mode in {"single", "mass"}:
        return mode
    if len(candidates) <= 1 or not interactive:
        return "single"
    print("\nНайдено несколько книг XLSM.")
    print("  1) single - выбрать одну книгу")
    print("  2) mass   - обработать несколько/все")
    answer = input("Выберите режим [1]: ").strip().lower()
    if answer in {"2", "mass", "m", "м"}:
        return "mass"
    return "single"


def _infer_files_dir_for_workbook(
    *,
    workbook_path: str,
    files_root: str,
    files_map: Dict[str, str],
    interactive: bool,
    prompt_always: bool,
) -> str:
    workbook_abs = str(Path(workbook_path).resolve())
    workbook_name = Path(workbook_abs).name
    workbook_stem = Path(workbook_abs).stem
    mapped_raw = files_map.get(workbook_abs) or files_map.get(workbook_name) or files_map.get(workbook_stem)

    default_dir = str(Path(files_root).resolve())
    if mapped_raw:
        mapped_path = Path(str(mapped_raw).strip())
        if not mapped_path.is_absolute():
            mapped_path = (Path(SCRIPT_DIR) / mapped_path).resolve()
        default_dir = str(mapped_path)
    else:
        root = Path(default_dir)
        if root.exists() and root.is_dir():
            for hint in _numeric_hints(workbook_name):
                for folder_name in (
                    hint,
                    f"part{hint}",
                    f"book{hint}",
                    "one" if hint == "1" else "",
                    "two" if hint == "2" else "",
                    "three" if hint == "3" else "",
                ):
                    if not folder_name:
                        continue
                    candidate = root / folder_name
                    if candidate.exists() and candidate.is_dir():
                        default_dir = str(candidate.resolve())
                        break

    if interactive and (prompt_always or not mapped_raw):
        chosen = _prompt_with_default(f"Files directory for {workbook_name}", default_dir, True)
        return str(Path(chosen).expanduser().resolve())
    return str(Path(default_dir).expanduser().resolve())


def _resolve_workbook_specs(args: argparse.Namespace, interactive: bool, logger) -> Tuple[List[WorkbookRunSpec], str]:
    explicit = parse_path_list(args.workbooks)
    if not explicit and args.workbook:
        explicit.append(args.workbook)
    explicit_raw = ";".join(explicit)

    candidates = discover_excel_files(SCRIPT_DIR, explicit_files=explicit_raw, pattern=EXCEL_INPUT_GLOB)
    if not candidates:
        candidates = discover_excel_files(SCRIPT_DIR, explicit_files="", pattern=EXCEL_INPUT_GLOB)
    if not candidates:
        raise FileNotFoundError("Книги XLSM не найдены")

    resolved_mode = _resolve_mode(args.mode, candidates, interactive)
    if resolved_mode == "single":
        chosen = _choose_single_workbook(candidates, interactive)
        selected = [chosen] if chosen else []
    else:
        selected = _choose_mass_workbooks(candidates, interactive)
    if not selected:
        raise RuntimeError("Не выбрана ни одна книга для миграции")

    files_root = str(Path(args.files_dir).expanduser().resolve())
    files_map = parse_key_value_mapping(args.files_map)

    specs: List[WorkbookRunSpec] = []
    for workbook_path in selected:
        workbook_abs = str(Path(workbook_path).expanduser().resolve())
        files_dir = _infer_files_dir_for_workbook(
            workbook_path=workbook_abs,
            files_root=files_root,
            files_map=files_map,
            interactive=interactive,
            prompt_always=bool(args.ask_files_always),
        )
        specs.append(WorkbookRunSpec(workbook_path=workbook_abs, files_dir=files_dir))

    logger.info("Выбрано книг: %s", len(specs))
    for idx, spec in enumerate(specs, start=1):
        logger.info("  %s) workbook=%s | files=%s", idx, spec.workbook_path, spec.files_dir)
    return specs, resolved_mode


def _resolve_runtime(args: argparse.Namespace) -> Tuple[str, str, str, str]:
    if args.profile == "custom":
        base_url = (args.base_url or BASE_URL).strip()
        jwt_url = (args.jwt_url or JWT_URL or (base_url.rstrip("/") + "/jwt/")).strip()
        ui_base_url = (args.ui_base_url or UI_BASE_URL or base_url).strip()
    else:
        profile = PROFILES[args.profile]
        base_url = (args.base_url or profile.base_url).strip()
        jwt_url = (args.jwt_url or profile.jwt_url).strip()
        ui_base_url = (args.ui_base_url or profile.ui_base_url).strip()

    if not base_url:
        raise ValueError("base_url is empty")
    if not jwt_url:
        jwt_url = base_url.rstrip("/") + "/jwt/"
    if not ui_base_url:
        ui_base_url = base_url
    return args.profile, base_url.rstrip("/"), jwt_url, ui_base_url.rstrip("/")


def _choose_resume_strategy(*, state: ResumeState, logger, user_logger, interactive: bool) -> str:
    if not state.enabled:
        return "continue"
    rows_count = state.rows_count()
    if rows_count <= 0:
        return "continue"

    run_info = state.get_run_info()
    status = str(run_info.get("status") or "").strip().lower()
    incomplete_statuses = {"running", "stopped", "aborted", "interrupted", "failed", "error", ""}
    if status not in incomplete_statuses:
        return "continue"

    last_checkpoint = run_info.get("lastCheckpoint") if isinstance(run_info.get("lastCheckpoint"), dict) else {}
    lines = [
        "Обнаружены незавершенные checkpoints прошлой миграции.",
        "",
        f"Статус прошлого запуска : {status or 'неизвестно (старый формат)'}",
        f"Старт запуска          : {_format_iso_for_console(run_info.get('startedAt'))}",
        f"Финиш запуска          : {_format_iso_for_console(run_info.get('finishedAt'))}",
        f"Профиль / стенд        : {run_info.get('profile') or '-'} / {run_info.get('baseUrl') or '-'}",
        f"Строк в checkpoint     : {rows_count}",
        f"Последняя точка        : {last_checkpoint.get('job') or '-'} / row={last_checkpoint.get('row') or '-'}",
        f"Последняя запись _id   : {last_checkpoint.get('_id') or '-'}",
    ]
    block = _console_block("RESUME: найдена незавершенная миграция", lines)
    logger.warning("%s", block)
    if user_logger:
        user_logger.info(block)

    if not interactive:
        logger.info("Интерактив отключен, автоматически продолжаем по checkpoint.")
        return "continue"

    prompt = "\n[RESUME] Действие: [П]родолжить / [С]бросить / [В]ыйти: "
    while True:
        try:
            raw = input(prompt)
        except EOFError:
            return "continue"
        choice = norm_str(raw).lower()
        if choice in {"", "п", "p", "продолжить", "continue", "resume"}:
            return "continue"
        if choice in {"с", "c", "сброс", "reset", "заново", "start"}:
            return "reset"
        if choice in {"в", "q", "quit", "выход", "exit", "abort"}:
            return "abort"


def _log_user_run_header(
    *,
    user_logger,
    profile: str,
    base_url: str,
    mode: str,
    interactive: bool,
    operator_mode: bool,
    state_file_path: str,
    success_log_path: str,
    fail_log_path: str,
) -> None:
    if user_logger is None:
        return
    lines = [
        "",
        "#" * 92,
        "===== СТАРТ МИГРАЦИИ НТО =====",
        "-" * 92,
        f"Профиль         : {profile}",
        f"Стенд           : {base_url}",
        f"Режим           : {mode}",
        f"Интерактивный   : {'да' if interactive else 'нет'}",
        f"Operator mode   : {'включен' if operator_mode else 'выключен'}",
        f"checkpoints     : {state_file_path}",
        f"Лог успехов     : {success_log_path or '-'}",
        f"Лог ошибок      : {fail_log_path or '-'}",
        "#" * 92,
    ]
    user_logger.info("\n".join(lines))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Миграция НТО из Excel напрямую в API")
    parser.add_argument("--profile", choices=["custom", "dev", "psi", "prod"], default="dev")
    parser.add_argument("--base-url", default="")
    parser.add_argument("--jwt-url", default="")
    parser.add_argument("--ui-base-url", default="")

    parser.add_argument("--mode", choices=["auto", "single", "mass"], default="auto")
    parser.add_argument("--workbook", default=str(Path(SCRIPT_DIR) / EXCEL_FILE_NAME), help="Путь к книге Excel (совместимость)")
    parser.add_argument("--workbooks", default="", help="Явный список книг (разделитель ';' или новая строка)")
    parser.add_argument("--files-dir", default=FILES_DIR, help="Папка с файлами для загрузки")
    parser.add_argument("--files-map", default="", help="Связка книга->папка файлов: 'book1.xlsm=dir1;book2.xlsm=dir2'")
    parser.add_argument("--ask-files-always", action="store_true", help="Всегда спрашивать папку файлов для каждой книги")

    parser.add_argument("--sheet", choices=["all", "mesto", "torgi"], default="all", help="Какие листы запускать")
    parser.add_argument("--limit", type=int, default=0, help="Ограничение по числу строк на лист")
    parser.add_argument("--dry-run", action="store_true", help="Только собрать payload без записи в API")
    parser.add_argument("--auth-only", action="store_true", help="Проверить авторизацию и завершить работу")
    parser.add_argument("--skip-auth", action="store_true", help="Пропустить авторизацию. Допустимо только с --dry-run")
    parser.add_argument("--no-auth", action="store_true", help=argparse.SUPPRESS)

    parser.add_argument("--operator-mode", action="store_true", help="На ошибке строки: retry/skip/abort")
    parser.add_argument("--no-prompt", action="store_true", help="Не запрашивать input, использовать значения из файлов")
    parser.add_argument("--no-interactive", action="store_true", help="Отключить интерактивный режим")
    parser.add_argument("--state-file", default=str(STATE_FILE), help="Путь к checkpoints JSON")
    parser.add_argument("--reset-state", action="store_true", help="Очистить checkpoints перед запуском")
    parser.add_argument("--resume", dest="resume", action="store_true", default=RESUME_BY_DEFAULT, help="Продолжать с checkpoints")
    parser.add_argument("--no-resume", dest="resume", action="store_false", help="Игнорировать checkpoints")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    warnings.filterwarnings("ignore", category=InsecureRequestWarning)

    logger = setup_logger()
    success_logger = setup_success_logger()
    fail_logger = setup_fail_logger()
    user_logger = setup_user_logger()

    interactive = not args.no_interactive and not args.no_prompt
    try:
        profile_name, runtime_base_url, runtime_jwt_url, runtime_ui_url = _resolve_runtime(args)
    except Exception as exc:
        logger.error("Ошибка runtime конфигурации: %s", exc)
        return 1
    set_runtime_urls(base_url=runtime_base_url, jwt_url=runtime_jwt_url, ui_base_url=runtime_ui_url)

    skip_auth = bool(args.skip_auth or args.no_auth)
    if skip_auth and not args.dry_run:
        logger.error("--skip-auth can be used only with --dry-run")
        return 1
    if args.auth_only and skip_auth:
        logger.error("--auth-only cannot be used together with --skip-auth")
        return 1

    try:
        workbook_specs, resolved_mode = _resolve_workbook_specs(args, interactive, logger)
    except Exception as exc:
        logger.error("Ошибка выбора книг/папок: %s", exc)
        return 1

    state_path = Path(args.state_file).expanduser().resolve()
    state_namespace = f"nto_migration:{profile_name}"
    state = ResumeState(path=state_path, namespace=state_namespace, enabled=True)
    if args.reset_state:
        state.reset_namespace()
        logger.info("State reset: %s", state_path)

    resume_enabled = bool(args.resume)
    if resume_enabled:
        strategy = _choose_resume_strategy(state=state, logger=logger, user_logger=user_logger, interactive=interactive)
        if strategy == "abort":
            logger.warning("Остановка по выбору оператора перед стартом")
            return 1
        if strategy == "reset":
            state.clear_rows()
            resume_enabled = False
            logger.info("Checkpoints очищены перед стартом")

    logger.info(
        "%s",
        _console_block(
            "NTO migration start",
            [
                f"Profile      : {profile_name}",
                f"Base URL     : {get_runtime_base_url()}",
                f"UI URL       : {get_runtime_ui_base_url()}",
                f"Mode         : {resolved_mode}",
                f"Sheet mode   : {args.sheet}",
                f"Dry run      : {args.dry_run}",
                f"Skip auth    : {skip_auth}",
                f"Resume       : {resume_enabled}",
                f"State file   : {state_path}",
                f"Interactive  : {interactive}",
                f"Operator mode: {args.operator_mode}",
                f"Workbook(s)  : {len(workbook_specs)}",
            ],
        ),
    )
    _log_user_run_header(
        user_logger=user_logger,
        profile=profile_name,
        base_url=get_runtime_base_url(),
        mode=resolved_mode,
        interactive=interactive,
        operator_mode=bool(args.operator_mode),
        state_file_path=str(state_path),
        success_log_path=getattr(success_logger, "log_path", ""),
        fail_log_path=getattr(fail_logger, "log_path", ""),
    )

    state.begin_run(
        profile=profile_name,
        baseUrl=get_runtime_base_url(),
        uiBaseUrl=get_runtime_ui_base_url(),
        mode=resolved_mode,
        sheet=args.sheet,
        dryRun=bool(args.dry_run),
        resume=resume_enabled,
        operatorMode=bool(args.operator_mode),
        workbooks=[spec.workbook_path for spec in workbook_specs],
        filesMap={spec.workbook_path: spec.files_dir for spec in workbook_specs},
    )

    session = None
    try:
        if not skip_auth:
            session = setup_session(logger, no_prompt=(args.no_prompt or args.no_interactive))
            logger.info("Авторизация выполнена успешно")

        if args.auth_only:
            logger.info("--auth-only: авторизация проверена, миграция не запускается")
            state.finish_run(
                status="completed",
                summary={
                    "authOnly": True,
                    "workbooks": len(workbook_specs),
                    "createdRows": 0,
                    "processedRows": 0,
                    "resumeSkips": 0,
                    "failedRows": 0,
                },
                clear_rows=True,
            )
            return 0
    except Exception as exc:
        logger.error("Авторизация не выполнена: %s", exc)
        state.finish_run(status="failed", summary={"reason": "auth_failed", "error": str(exc)}, clear_rows=False)
        return 1

    rollback_items: List[Dict[str, Any]] = []

    processed_rows = 0
    created_rows = 0
    resumed_skips = 0
    failed_rows = 0
    stopped = False

    try:
        for spec in workbook_specs:
            workbook_path = Path(spec.workbook_path).expanduser().resolve()
            files_dir = Path(spec.files_dir).expanduser().resolve()
            base_dir = Path(SCRIPT_DIR).resolve()

            if not workbook_path.exists():
                raise FileNotFoundError(f"Workbook not found: {workbook_path}")

            logger.info("%s", _console_block("Workbook", [f"Path: {workbook_path}", f"Files: {files_dir}"]))
            workbook = load_workbook(workbook_path, data_only=True)

            for sheet_name, kind in _iter_selected_sheets(args.sheet):
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"Лист не найден: {sheet_name} (в книге {workbook_path.name})")
                ws = workbook[sheet_name]
                logger.info("Обработка листа: %s | книга: %s", sheet_name, workbook_path.name)

                for row_idx, row_data in iter_vba_rows(ws, row_limit=max(0, int(args.limit or 0))):
                    row_key = state.get(str(workbook_path), sheet_name, row_idx)
                    if resume_enabled and isinstance(row_key, dict):
                        resumed_skips += 1
                        logger.info(
                            "[RESUME][SKIP] workbook=%s sheet=%s row=%s _id=%s",
                            workbook_path.name,
                            sheet_name,
                            row_idx,
                            row_key.get("_id"),
                        )
                        continue

                    processed_rows += 1
                    while True:
                        try:
                            if kind == "mesto":
                                result = _process_mesto_row(
                                    row_idx=row_idx,
                                    row_data=row_data,
                                    session=session,
                                    logger=logger,
                                    success_logger=success_logger,
                                    rollback_items=rollback_items,
                                    files_dir=files_dir,
                                    base_dir=base_dir,
                                    dry_run=args.dry_run,
                                )
                            else:
                                result = _process_torgi_row(
                                    row_idx=row_idx,
                                    row_data=row_data,
                                    session=session,
                                    logger=logger,
                                    success_logger=success_logger,
                                    rollback_items=rollback_items,
                                    files_dir=files_dir,
                                    base_dir=base_dir,
                                    dry_run=args.dry_run,
                                )

                            if not args.dry_run and isinstance(result, dict) and result.get("_id"):
                                created_rows += 1
                                state.mark_success(
                                    workbook_path=str(workbook_path),
                                    job_name=sheet_name,
                                    row_idx=row_idx,
                                    collection=str(result.get("collection") or ""),
                                    main_id=str(result.get("_id") or ""),
                                    guid=str(result.get("guid") or ""),
                                    had_errors=False,
                                    error_count=0,
                                )
                            break
                        except Exception as exc:
                            failed_rows += 1
                            error_payload = {
                                "workbook": str(workbook_path),
                                "sheet": sheet_name,
                                "row": row_idx,
                                "error": str(exc),
                                "traceback": traceback.format_exc(limit=30),
                            }
                            fail_logger.info(json.dumps(error_payload, ensure_ascii=False))
                            logger.exception("[FAIL][%s][ROW:%s][WB:%s]", sheet_name, row_idx, workbook_path.name)
                            user_logger.info(
                                "ROW_ERROR | workbook=%s | sheet=%s | row=%s | error=%s",
                                workbook_path.name,
                                sheet_name,
                                row_idx,
                                str(exc),
                            )

                            if args.operator_mode and interactive:
                                action = _ask_error_action()
                                if action == "retry":
                                    continue
                                if action == "skip":
                                    logger.warning("[ROW][SKIP] workbook=%s sheet=%s row=%s", workbook_path.name, sheet_name, row_idx)
                                    break
                                stopped = True
                                break

                            stopped = True
                            break

                    if stopped:
                        logger.warning("Migration stopped on workbook=%s sheet=%s row=%s", workbook_path.name, sheet_name, row_idx)
                        break

                if stopped:
                    break

            if stopped:
                break

    except KeyboardInterrupt:
        stopped = True
        logger.warning("Остановка по Ctrl+C")
    except Exception as exc:
        stopped = True
        logger.error("Критическая ошибка выполнения: %s", exc)
        logger.debug(traceback.format_exc())
    finally:
        summary = {
            "processedRows": processed_rows,
            "createdRows": created_rows,
            "resumeSkips": resumed_skips,
            "failedRows": failed_rows,
            "rollbackItems": len(rollback_items),
            "workbooks": len(workbook_specs),
            "profile": profile_name,
            "baseUrl": get_runtime_base_url(),
            "mode": resolved_mode,
            "dryRun": bool(args.dry_run),
            "skipAuth": bool(skip_auth),
        }
        if stopped:
            state.finish_run(status="stopped", summary=summary, clear_rows=False)
        elif failed_rows > 0:
            state.finish_run(status="failed", summary=summary, clear_rows=False)
        else:
            state.finish_run(status="completed", summary=summary, clear_rows=True)

        user_logger.info("FINISH | status=%s | summary=%s", "stopped" if stopped else ("failed" if failed_rows > 0 else "completed"), json.dumps(summary, ensure_ascii=False))

    if rollback_items:
        _write_rollback_body(rollback_items)

    logger.info(
        "%s",
        _console_block(
            "NTO migration finish",
            [
                f"Created rows   : {created_rows}",
                f"Processed rows : {processed_rows}",
                f"Resume skips   : {resumed_skips}",
                f"Failed rows    : {failed_rows}",
                f"Rollback items : {len(rollback_items)}",
                f"Status         : {'stopped' if stopped else ('failed' if failed_rows > 0 else 'completed')}",
            ],
        ),
    )

    if stopped:
        return 1
    if failed_rows > 0:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
