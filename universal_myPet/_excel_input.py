import glob
import math
import os
import re
from datetime import date, datetime, time, timedelta

from _config import EXCEL_DATA_START_ROW


def _parse_path_list(raw):
    out = []
    text = str(raw or "").strip()
    if not text:
        return out
    for part in re.split(r"[;\n]+", text):
        token = part.strip().strip('"').strip("'").strip()
        if token:
            out.append(token)
    return out


def discover_excel_files(script_dir, explicit_files="", pattern="*.xlsm"):
    explicit = _parse_path_list(explicit_files)
    if explicit:
        out = []
        for raw in explicit:
            candidate = raw
            if not os.path.isabs(candidate):
                candidate = os.path.join(script_dir, candidate)
            candidate = os.path.abspath(candidate)
            if os.path.isfile(candidate) and not os.path.basename(candidate).startswith("~$"):
                out.append(candidate)
        return sorted(list(dict.fromkeys(out)), key=lambda p: os.path.basename(p).lower())

    files = [
        os.path.abspath(p)
        for p in glob.glob(os.path.join(script_dir, pattern))
        if os.path.isfile(p) and not os.path.basename(p).startswith("~$")
    ]
    files.sort(key=lambda p: os.path.basename(p).lower())
    return files


def discover_excel_file(script_dir, explicit_file="", pattern="*.xlsm"):
    files = discover_excel_files(script_dir, explicit_files=explicit_file, pattern=pattern)
    return files[0] if files else None


def _is_empty(value):
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _to_scalar(value):
    if _is_empty(value):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return value
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    return value


def _digits_only(value):
    s = str(_to_scalar(value) or "")
    d = re.sub(r"\D+", "", s)
    return d or None


def _excel_serial_to_date(serial):
    try:
        num = float(serial)
    except Exception:
        return None
    if num < 20000 or num > 80000:
        return None
    base = datetime(1899, 12, 30)
    return (base + timedelta(days=num)).strftime("%Y-%m-%d")


def _to_date_text(value):
    if _is_empty(value):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        from_serial = _excel_serial_to_date(value)
        if from_serial:
            return from_serial
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    s = str(value).strip()
    if not s:
        return None
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    if re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", s):
        dd, mm, yyyy = s.split(".")
        return f"{yyyy}-{mm}-{dd}"
    from_serial = _excel_serial_to_date(s.replace(",", "."))
    if from_serial:
        return from_serial
    return s


def _to_time_text(value):
    if _is_empty(value):
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)):
        try:
            n = float(value)
            if 0 <= n < 2:
                total_minutes = int(round((n % 1) * 24 * 60))
                hh = (total_minutes // 60) % 24
                mm = total_minutes % 60
                return f"{hh:02d}:{mm:02d}"
        except Exception:
            pass
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    s = str(value).strip()
    if not s:
        return None
    m = re.match(r"^\s*(\d{1,2}):(\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return s


def _split_file_tokens(value):
    s = str(_to_scalar(value) or "").strip()
    if not s:
        return []
    s = s.replace("\r\n", "\n").replace("\r", "\n").replace("|", ";").replace(",", ";")
    s = s.replace("\n", ";")
    out = []
    for part in s.split(";"):
        token = part.strip().strip('"').strip("'").strip()
        if token:
            out.append(token)
    return out


def _put_if(target, key, value):
    if not _is_empty(value):
        target[key] = value


def _set_file_fields(target, single_key, value, filename_key=None, list_key=None):
    files = _split_file_tokens(value)
    if not files:
        return
    target[single_key] = files[0]
    if filename_key:
        target[filename_key] = files[0]
    if list_key and len(files) > 1:
        target[list_key] = [{"fileName": x} for x in files]


def _row_has_any(ws, row_index, col_from, col_to):
    for col in range(col_from, col_to + 1):
        if not _is_empty(ws.cell(row_index, col).value):
            return True
    return False


def _parse_catch_rows(ws):
    rows = []
    max_col = min(ws.max_column, 151)
    for r in range(EXCEL_DATA_START_ROW, ws.max_row + 1):
        if not _row_has_any(ws, r, 2, max_col):
            continue

        order = {}
        _put_if(order, "region", _to_scalar(ws.cell(r, 2).value))
        _put_if(order, "authorizedOrgName", _to_scalar(ws.cell(r, 3).value))
        _put_if(order, "ogrn", _digits_only(ws.cell(r, 4).value))
        _put_if(order, "inn", _digits_only(ws.cell(r, 5).value))
        _put_if(order, "catchOrgName", _to_scalar(ws.cell(r, 6).value))
        _put_if(order, "catchOrgOgrn", _digits_only(ws.cell(r, 7).value))
        _put_if(order, "catchOrgInn", _digits_only(ws.cell(r, 8).value))
        _put_if(order, "orderNumber", _to_scalar(ws.cell(r, 9).value))
        _put_if(order, "priority", _to_scalar(ws.cell(r, 10).value))
        _put_if(order, "catchDays", _to_scalar(ws.cell(r, 11).value))

        animals = []
        for start in range(12, max_col + 1, 14):
            if start + 13 > max_col:
                break
            vals = [ws.cell(r, start + i).value for i in range(14)]
            if all(_is_empty(v) for v in vals):
                continue

            item = {}
            _put_if(item, "number", _to_scalar(vals[0]))
            _put_if(item, "kind", _to_scalar(vals[1]))
            _put_if(item, "color", _to_scalar(vals[2]))
            _put_if(item, "size", _to_scalar(vals[3]))
            _put_if(item, "unmotivatedAggression", _to_scalar(vals[4]))
            _put_if(item, "aggressionDescription", _to_scalar(vals[5]))
            _put_if(item, "clip", _to_scalar(vals[6]))
            _put_if(item, "clipColor", _to_scalar(vals[7]))
            _put_if(item, "extraInfo", _to_scalar(vals[8]))
            _put_if(item, "locationAddress", _to_scalar(vals[9]))
            _put_if(item, "locationLandmark", _to_scalar(vals[10]))
            _set_file_fields(item, "photo", vals[11], filename_key="photoFileName")
            _put_if(item, "status", _to_scalar(vals[12]))
            _put_if(item, "note", _to_scalar(vals[13]))

            if item:
                animals.append(item)

        if order or animals:
            rows.append({"orderInfo": order, "animals": animals})
    return rows


def _parse_stray_rows(ws):
    rows = []
    max_col = min(ws.max_column, 36)
    for r in range(EXCEL_DATA_START_ROW, ws.max_row + 1):
        if not _row_has_any(ws, r, 2, max_col):
            continue

        row = {}
        _put_if(row, "region", _to_scalar(ws.cell(r, 2).value))
        _put_if(row, "authorizedOrgName", _to_scalar(ws.cell(r, 3).value))
        _put_if(row, "ogrn", _digits_only(ws.cell(r, 4).value))
        _put_if(row, "inn", _digits_only(ws.cell(r, 5).value))
        _put_if(row, "catchOrgName", _to_scalar(ws.cell(r, 6).value))
        _put_if(row, "catchOrgOgrn", _digits_only(ws.cell(r, 7).value))
        _put_if(row, "catchOrgInn", _digits_only(ws.cell(r, 8).value))
        _put_if(row, "animalNumber", _to_scalar(ws.cell(r, 9).value))
        _put_if(row, "type", _to_scalar(ws.cell(r, 10).value))
        _put_if(row, "sex", _to_scalar(ws.cell(r, 11).value))
        _put_if(row, "coloration", _to_scalar(ws.cell(r, 12).value))
        _put_if(row, "size", _to_scalar(ws.cell(r, 13).value))
        _put_if(row, "unmotivatedAggression", _to_scalar(ws.cell(r, 14).value))
        _put_if(row, "aggressionDescription", _to_scalar(ws.cell(r, 15).value))
        _put_if(row, "clip", _to_scalar(ws.cell(r, 16).value))
        _put_if(row, "clipColor", _to_scalar(ws.cell(r, 17).value))
        _put_if(row, "additionalInfo", _to_scalar(ws.cell(r, 18).value))
        _put_if(row, "locationAddress", _to_scalar(ws.cell(r, 19).value))
        _put_if(row, "locationLandmark", _to_scalar(ws.cell(r, 20).value))
        _set_file_fields(row, "photo", ws.cell(r, 21).value, filename_key="photoFileName")
        _put_if(row, "animalStatus", _to_scalar(ws.cell(r, 22).value))
        _set_file_fields(row, "note", ws.cell(r, 23).value, filename_key="noteFileName")
        _put_if(row, "orderNumber", _to_scalar(ws.cell(r, 24).value))
        _put_if(row, "municipalContractNumber", _to_scalar(ws.cell(r, 25).value))
        _put_if(row, "municipalContractDate", _to_date_text(ws.cell(r, 26).value))
        _put_if(row, "catchStartDate", _to_date_text(ws.cell(r, 27).value))
        _put_if(row, "catchStartTime", _to_time_text(ws.cell(r, 28).value))
        _put_if(row, "catchEndDate", _to_date_text(ws.cell(r, 29).value))
        _put_if(row, "catchEndTime", _to_time_text(ws.cell(r, 30).value))
        _put_if(row, "catchAddress", _to_scalar(ws.cell(r, 31).value))
        _set_file_fields(row, "catchVideo", ws.cell(r, 32).value, filename_key="catchVideoFileName")
        _put_if(row, "catcherFIO", _to_scalar(ws.cell(r, 33).value))
        _put_if(row, "catchActNumber", _to_scalar(ws.cell(r, 34).value))
        _put_if(row, "catchActDate", _to_date_text(ws.cell(r, 35).value))
        _set_file_fields(row, "catchAct", ws.cell(r, 36).value, filename_key="catchActFileName")

        if row:
            rows.append(row)
    return rows


def _build_dworm_item(ws, r, start, with_type=False):
    item = {}
    if with_type:
        _put_if(item, "type", _to_scalar(ws.cell(r, start).value))
        _put_if(item, "drugName", _to_scalar(ws.cell(r, start + 1).value))
        _put_if(item, "dosage", _to_scalar(ws.cell(r, start + 2).value))
        _put_if(item, "date", _to_date_text(ws.cell(r, start + 3).value))
        _put_if(item, "employeeFIO", _to_scalar(ws.cell(r, start + 4).value))
        _put_if(item, "employeePosition", _to_scalar(ws.cell(r, start + 5).value))
        _put_if(item, "actNumber", _to_scalar(ws.cell(r, start + 6).value))
        _set_file_fields(item, "dewormAct", ws.cell(r, start + 7).value, filename_key="dewormActFileName", list_key="dewormActs")
    else:
        _put_if(item, "drugName", _to_scalar(ws.cell(r, start).value))
        _put_if(item, "dosage", _to_scalar(ws.cell(r, start + 1).value))
        _put_if(item, "date", _to_date_text(ws.cell(r, start + 2).value))
        _put_if(item, "employeeFIO", _to_scalar(ws.cell(r, start + 3).value))
        _put_if(item, "employeePosition", _to_scalar(ws.cell(r, start + 4).value))
        _put_if(item, "actNumber", _to_scalar(ws.cell(r, start + 5).value))
        _set_file_fields(item, "dewormAct", ws.cell(r, start + 6).value, filename_key="dewormActFileName", list_key="dewormActs")
    return item if item else None


def _build_disinsection_item(ws, r, start):
    item = {}
    _put_if(item, "drugName", _to_scalar(ws.cell(r, start).value))
    _put_if(item, "date", _to_date_text(ws.cell(r, start + 1).value))
    _put_if(item, "employeeFIO", _to_scalar(ws.cell(r, start + 2).value))
    _put_if(item, "employeePosition", _to_scalar(ws.cell(r, start + 3).value))
    _put_if(item, "actNumber", _to_scalar(ws.cell(r, start + 4).value))
    _set_file_fields(item, "disinsectionAct", ws.cell(r, start + 5).value, filename_key="disinsectionActFileName", list_key="disinsectionActs")
    return item if item else None


def _build_vaccination_item(ws, r, start):
    item = {}
    _put_if(item, "drugName", _to_scalar(ws.cell(r, start).value))
    _put_if(item, "date", _to_date_text(ws.cell(r, start + 1).value))
    _put_if(item, "series", _to_scalar(ws.cell(r, start + 2).value))
    _put_if(item, "dosage", _to_scalar(ws.cell(r, start + 3).value))
    _put_if(item, "employeeFIO", _to_scalar(ws.cell(r, start + 4).value))
    _put_if(item, "employeePosition", _to_scalar(ws.cell(r, start + 5).value))
    _put_if(item, "actNumber", _to_scalar(ws.cell(r, start + 6).value))
    _set_file_fields(item, "vaccinationAct", ws.cell(r, start + 7).value, filename_key="vaccinationActFileName", list_key="vaccinationActs")
    return item if item else None


def _build_card_row(ws, r):
    row = {}
    _put_if(row, "region", _to_scalar(ws.cell(r, 2).value))
    _put_if(row, "authorizedOrgName", _to_scalar(ws.cell(r, 3).value))
    _put_if(row, "ogrn", _digits_only(ws.cell(r, 4).value))
    _put_if(row, "inn", _digits_only(ws.cell(r, 5).value))
    _put_if(row, "shelterName", _to_scalar(ws.cell(r, 6).value))
    _put_if(row, "shelterOGRN", _digits_only(ws.cell(r, 7).value))
    _put_if(row, "shelterINN", _digits_only(ws.cell(r, 8).value))
    _put_if(row, "cardNumber", _to_scalar(ws.cell(r, 9).value))
    _put_if(row, "type", _to_scalar(ws.cell(r, 10).value))
    _put_if(row, "sex", _to_scalar(ws.cell(r, 11).value))
    _put_if(row, "breed", _to_scalar(ws.cell(r, 12).value))
    _put_if(row, "coloration", _to_scalar(ws.cell(r, 13).value))
    _put_if(row, "size", _to_scalar(ws.cell(r, 14).value))
    _put_if(row, "nickname", _to_scalar(ws.cell(r, 15).value))
    _put_if(row, "fur", _to_scalar(ws.cell(r, 16).value))
    _put_if(row, "ears", _to_scalar(ws.cell(r, 17).value))
    _put_if(row, "tail", _to_scalar(ws.cell(r, 18).value))
    _put_if(row, "specialMarks", _to_scalar(ws.cell(r, 19).value))
    _put_if(row, "age", _to_scalar(ws.cell(r, 20).value))
    _put_if(row, "weight", _to_scalar(ws.cell(r, 21).value))
    _put_if(row, "temperature", _to_scalar(ws.cell(r, 22).value))
    _put_if(row, "injuriesInfo", _to_scalar(ws.cell(r, 23).value))
    _put_if(row, "numMarker", _to_scalar(ws.cell(r, 24).value))
    _put_if(row, "methodMarker", _to_scalar(ws.cell(r, 25).value))
    _put_if(row, "cageNumber", _to_scalar(ws.cell(r, 26).value))
    _set_file_fields(row, "photo", ws.cell(r, 27).value, filename_key="photoFileName", list_key="photos")
    _put_if(row, "animalStatus", _to_scalar(ws.cell(r, 28).value))
    _put_if(row, "releaseFromShelterDate", _to_date_text(ws.cell(r, 29).value))
    _put_if(row, "releaseFromPVSDate", _to_date_text(ws.cell(r, 30).value))
    _put_if(row, "quarantineUntilDate", _to_date_text(ws.cell(r, 31).value))

    if row.get("numMarker") or row.get("methodMarker"):
        row["identityMark"] = {
            "number": row.get("numMarker"),
            "method": row.get("methodMarker"),
        }

    dewormings = []
    first = _build_dworm_item(ws, r, 32, with_type=True)
    if first:
        dewormings.append(first)
    for start in (40, 47, 54, 61, 68, 75, 82, 89, 96):
        item = _build_dworm_item(ws, r, start, with_type=False)
        if item:
            dewormings.append(item)
    if dewormings:
        row["dewormings"] = dewormings

    disinsections = []
    for start in (104, 110, 116, 122, 128, 134, 140, 146, 152, 158):
        item = _build_disinsection_item(ws, r, start)
        if item:
            disinsections.append(item)
    if disinsections:
        row["disinsections"] = disinsections

    vaccinations = []
    for start in (165, 173, 181, 189, 197, 205, 213, 221, 229, 237):
        item = _build_vaccination_item(ws, r, start)
        if item:
            vaccinations.append(item)
    if vaccinations:
        row["vaccinations"] = vaccinations

    ster = {}
    _put_if(ster, "drugName", _to_scalar(ws.cell(r, 246).value))
    _put_if(ster, "date", _to_date_text(ws.cell(r, 247).value))
    _put_if(ster, "dose", _to_scalar(ws.cell(r, 248).value))
    _put_if(ster, "employeeFIO", _to_scalar(ws.cell(r, 249).value))
    _put_if(ster, "employeePosition", _to_scalar(ws.cell(r, 250).value))
    _put_if(ster, "actNumber", _to_scalar(ws.cell(r, 251).value))
    _set_file_fields(ster, "sterilizationAct", ws.cell(r, 252).value, filename_key="sterilizationActFileName", list_key="sterilizationActs")
    if ster:
        row["sterilizations"] = [ster]

    exam = {}
    _put_if(exam, "foodReactionPresence", _to_scalar(ws.cell(r, 254).value))
    _put_if(exam, "foodReactionOffer", _to_scalar(ws.cell(r, 255).value))
    _put_if(exam, "loudSoundReaction", _to_scalar(ws.cell(r, 256).value))
    _put_if(exam, "commissionDecision", _to_scalar(ws.cell(r, 257).value))
    _put_if(exam, "commissionMember257", _to_scalar(ws.cell(r, 258).value))
    _put_if(exam, "commissionMember258", _to_scalar(ws.cell(r, 259).value))
    _put_if(exam, "commissionMember259", _to_scalar(ws.cell(r, 260).value))
    _put_if(exam, "commissionMember260", _to_scalar(ws.cell(r, 261).value))
    _put_if(exam, "commissionMember261", _to_scalar(ws.cell(r, 262).value))
    _put_if(exam, "actAuthor", _to_scalar(ws.cell(r, 263).value))
    _put_if(exam, "date", _to_date_text(ws.cell(r, 264).value))
    _put_if(exam, "actNumber", _to_scalar(ws.cell(r, 265).value))
    _set_file_fields(exam, "actFile", ws.cell(r, 266).value, filename_key="actFileFileName", list_key="actFiles")
    if exam:
        row["examination"] = exam

    euth = {}
    _put_if(euth, "reason", _to_scalar(ws.cell(r, 268).value))
    _put_if(euth, "date", _to_date_text(ws.cell(r, 269).value))
    _put_if(euth, "time", _to_time_text(ws.cell(r, 270).value))
    _put_if(euth, "method", _to_scalar(ws.cell(r, 271).value))
    _put_if(euth, "drugName", _to_scalar(ws.cell(r, 272).value))
    _put_if(euth, "dosage", _to_scalar(ws.cell(r, 273).value))
    _put_if(euth, "employeeFIO", _to_scalar(ws.cell(r, 274).value))
    _put_if(euth, "employeePosition", _to_scalar(ws.cell(r, 275).value))
    _put_if(euth, "actNumber", _to_scalar(ws.cell(r, 276).value))
    _set_file_fields(euth, "actFile", ws.cell(r, 277).value, filename_key="actFileFileName", list_key="actFiles")
    if euth:
        row["euthanasia"] = euth

    util = {}
    _put_if(util, "date", _to_date_text(ws.cell(r, 279).value))
    _put_if(util, "basis", _to_scalar(ws.cell(r, 280).value))
    _put_if(util, "method", _to_scalar(ws.cell(r, 281).value))
    _put_if(util, "employeeFIO", _to_scalar(ws.cell(r, 282).value))
    _put_if(util, "employeePosition", _to_scalar(ws.cell(r, 283).value))
    _put_if(util, "actNumber", _to_scalar(ws.cell(r, 284).value))
    _set_file_fields(util, "actFile", ws.cell(r, 285).value, filename_key="actFileFileName", list_key="actFiles")
    if util:
        row["utilization"] = util

    mark = {}
    _put_if(mark, "number", _to_scalar(ws.cell(r, 287).value))
    _put_if(mark, "method", _to_scalar(ws.cell(r, 288).value))
    _put_if(mark, "place", _to_scalar(ws.cell(r, 289).value))
    _put_if(mark, "date", _to_date_text(ws.cell(r, 290).value))
    _put_if(mark, "employeeFIO", _to_scalar(ws.cell(r, 291).value))
    _put_if(mark, "employeePosition", _to_scalar(ws.cell(r, 292).value))
    if mark:
        row["markingEvents"] = [mark]

    other_events = []
    for start in (294, 301, 308, 315, 322, 329, 336, 343, 350, 357):
        evt = {}
        _put_if(evt, "name", _to_scalar(ws.cell(r, start).value))
        _put_if(evt, "description", _to_scalar(ws.cell(r, start + 1).value))
        _put_if(evt, "date", _to_date_text(ws.cell(r, start + 2).value))
        _put_if(evt, "employeeFIO", _to_scalar(ws.cell(r, start + 3).value))
        _put_if(evt, "employeePosition", _to_scalar(ws.cell(r, start + 4).value))
        _put_if(evt, "documentNumber", _to_scalar(ws.cell(r, start + 5).value))
        _set_file_fields(evt, "otherEventDocument", ws.cell(r, start + 6).value, filename_key="otherEventDocumentFileName", list_key="otherEventDocuments")
        if evt:
            other_events.append(evt)
    if other_events:
        row["otherEvents"] = other_events

    release = {}
    _put_if(release, "actName", _to_scalar(ws.cell(r, 364).value))
    _put_if(release, "actNumber", _to_scalar(ws.cell(r, 365).value))
    _put_if(release, "actDate", _to_date_text(ws.cell(r, 366).value))
    _put_if(release, "shelterName", _to_scalar(ws.cell(r, 367).value))
    _put_if(release, "shelterAddress", _to_scalar(ws.cell(r, 368).value))
    _put_if(release, "shelterINN", _digits_only(ws.cell(r, 369).value))
    _put_if(release, "shelterOGRN", _digits_only(ws.cell(r, 370).value))
    _put_if(release, "pvsName", _to_scalar(ws.cell(r, 371).value))
    _put_if(release, "pvsAddress", _to_scalar(ws.cell(r, 372).value))
    _put_if(release, "pvsINN", _digits_only(ws.cell(r, 373).value))
    _put_if(release, "pvsOGRN", _digits_only(ws.cell(r, 374).value))
    _put_if(release, "catcherFIO", _to_scalar(ws.cell(r, 375).value))
    _put_if(release, "releaseAddress", _to_scalar(ws.cell(r, 376).value))
    _set_file_fields(release, "actFile", ws.cell(r, 377).value, filename_key="actFileFileName", list_key="actFiles")
    if release:
        row["releaseInfo"] = release

    transfer = {}
    _put_if(transfer, "actName", _to_scalar(ws.cell(r, 378).value))
    _put_if(transfer, "actNumber", _to_scalar(ws.cell(r, 379).value))
    _put_if(transfer, "transferDate", _to_date_text(ws.cell(r, 380).value))
    _put_if(transfer, "shelterName", _to_scalar(ws.cell(r, 381).value))
    _put_if(transfer, "shelterAddress", _to_scalar(ws.cell(r, 382).value))
    _put_if(transfer, "shelterINN", _digits_only(ws.cell(r, 383).value))
    _put_if(transfer, "shelterOGRN", _digits_only(ws.cell(r, 384).value))
    _put_if(transfer, "pvsName", _to_scalar(ws.cell(r, 385).value))
    _put_if(transfer, "pvsAddress", _to_scalar(ws.cell(r, 386).value))
    _put_if(transfer, "pvsINN", _digits_only(ws.cell(r, 387).value))
    _put_if(transfer, "pvsOGRN", _digits_only(ws.cell(r, 388).value))
    _put_if(transfer, "newOwnerFIO", _to_scalar(ws.cell(r, 389).value))
    _put_if(transfer, "newOwnerAddress", _to_scalar(ws.cell(r, 390).value))
    _put_if(transfer, "idSeries", _to_scalar(ws.cell(r, 391).value))
    _put_if(transfer, "idNumber", _to_scalar(ws.cell(r, 392).value))
    _put_if(transfer, "idDeptCode", _to_scalar(ws.cell(r, 393).value))
    _put_if(transfer, "idIssueDate", _to_date_text(ws.cell(r, 394).value))
    _put_if(transfer, "idIssuedBy", _to_scalar(ws.cell(r, 395).value))
    _set_file_fields(transfer, "actFile", ws.cell(r, 396).value, filename_key="actFileFileName", list_key="actFiles")
    if transfer:
        row["transferToOwner"] = transfer

    death = {}
    _put_if(death, "actName", _to_scalar(ws.cell(r, 397).value))
    _put_if(death, "actNumber", _to_scalar(ws.cell(r, 398).value))
    _put_if(death, "actDate", _to_date_text(ws.cell(r, 399).value))
    _put_if(death, "deathDate", _to_date_text(ws.cell(r, 400).value))
    _put_if(death, "shelterName", _to_scalar(ws.cell(r, 401).value))
    _put_if(death, "shelterAddress", _to_scalar(ws.cell(r, 402).value))
    _put_if(death, "shelterINN", _digits_only(ws.cell(r, 403).value))
    _put_if(death, "shelterOGRN", _digits_only(ws.cell(r, 404).value))
    _put_if(death, "pvsName", _to_scalar(ws.cell(r, 405).value))
    _put_if(death, "pvsAddress", _to_scalar(ws.cell(r, 406).value))
    _put_if(death, "pvsINN", _digits_only(ws.cell(r, 407).value))
    _put_if(death, "pvsOGRN", _digits_only(ws.cell(r, 408).value))
    _set_file_fields(death, "actFile", ws.cell(r, 409).value, filename_key="actFileFileName", list_key="actFiles")
    if death:
        row["deathInfo"] = death

    hand_c = {}
    _put_if(hand_c, "actName", _to_scalar(ws.cell(r, 410).value))
    _put_if(hand_c, "actNumber", _to_scalar(ws.cell(r, 411).value))
    _put_if(hand_c, "orderNumber", _to_scalar(ws.cell(r, 412).value))
    _put_if(hand_c, "orderCreateDate", _to_date_text(ws.cell(r, 413).value))
    _put_if(hand_c, "catcherFIO", _to_scalar(ws.cell(r, 414).value))
    _put_if(hand_c, "catcherPhone", _to_scalar(ws.cell(r, 415).value))
    _put_if(hand_c, "shelterName", _to_scalar(ws.cell(r, 416).value))
    _put_if(hand_c, "shelterAddress", _to_scalar(ws.cell(r, 417).value))
    _put_if(hand_c, "shelterPhone", _to_scalar(ws.cell(r, 418).value))
    _put_if(hand_c, "pvsName", _to_scalar(ws.cell(r, 419).value))
    _put_if(hand_c, "pvsAddress", _to_scalar(ws.cell(r, 420).value))
    _put_if(hand_c, "pvsPhone", _to_scalar(ws.cell(r, 421).value))
    _set_file_fields(hand_c, "actFile", ws.cell(r, 422).value, filename_key="actFileFileName", list_key="actFiles")
    if hand_c:
        row["handoverWithCatcher"] = hand_c

    hand_s = {}
    _put_if(hand_s, "actName", _to_scalar(ws.cell(r, 423).value))
    _put_if(hand_s, "actNumber", _to_scalar(ws.cell(r, 424).value))
    _put_if(hand_s, "actDate", _to_date_text(ws.cell(r, 425).value))
    _put_if(hand_s, "pvsName", _to_scalar(ws.cell(r, 426).value))
    _put_if(hand_s, "shelterName", _to_scalar(ws.cell(r, 427).value))
    _set_file_fields(hand_s, "actFile", ws.cell(r, 428).value, filename_key="actFileFileName", list_key="actFiles")
    if hand_s:
        row["handoverWithShelter"] = hand_s

    return row


def _parse_card_rows(ws):
    rows = []
    max_col = min(ws.max_column, 428)
    for r in range(EXCEL_DATA_START_ROW, ws.max_row + 1):
        if not _row_has_any(ws, r, 2, max_col):
            continue
        row = _build_card_row(ws, r)
        if row:
            rows.append(row)
    return rows


def load_rows_from_excel(excel_path, logger):
    def _patch_numpy_compat_for_openpyxl26():
        try:
            import numpy as np
        except Exception:
            return
        aliases = {
            "float": float,
            "int": int,
            "bool": bool,
            "object": object,
            "str": str,
        }
        namespace = getattr(np, "__dict__", {})
        for name, value in aliases.items():
            if name not in namespace:
                try:
                    setattr(np, name, value)
                except Exception:
                    pass

    try:
        _patch_numpy_compat_for_openpyxl26()
        import openpyxl
    except Exception as exc:
        raise RuntimeError("openpyxl is required for direct Excel input. Install: pip install openpyxl") from exc

    with open(str(excel_path), "rb") as fp:
        wb = openpyxl.load_workbook(fp, data_only=True, keep_vba=True)
    if len(wb.worksheets) < 4:
        raise RuntimeError("Workbook must contain at least 4 sheets (title + 3 registries)")

    catch_ws = wb.worksheets[1]
    stray_ws = wb.worksheets[2]
    card_ws = wb.worksheets[3]

    logger.info("[EXCEL] workbook=%s", os.path.basename(excel_path))
    logger.info("[EXCEL] sheets: catch='%s', stray='%s', card='%s'", catch_ws.title, stray_ws.title, card_ws.title)

    catch_rows = _parse_catch_rows(catch_ws)
    stray_rows = _parse_stray_rows(stray_ws)
    card_rows = _parse_card_rows(card_ws)

    logger.info("[EXCEL] parsed rows: catch=%s stray=%s card=%s", len(catch_rows), len(stray_rows), len(card_rows))
    return {"catch": catch_rows, "stray": stray_rows, "card": card_rows}
